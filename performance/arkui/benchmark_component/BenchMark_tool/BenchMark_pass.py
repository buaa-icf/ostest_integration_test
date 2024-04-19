# Copyright (c) 2023 Huawei Device Co., Ltd.
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import openpyxl

def read_excel_file(filename):
    # 打开xlsx，得到workbook对象
    wb = openpyxl.load_workbook(filename)
    sheet_names = wb.sheetnames
    for name in sheet_names:
        print('sheet_name= ', name)
        if '基线数据' != name:
            continue
        ws = wb[name]
        for i in range(ws.max_row):
            if i==0:
               continue
            # 每一行的第7列
            cell_s = ws.cell(row=i+1, column=7)
            print('cell_s.value= ', cell_s.value)
            # 每一行的第8列
            cell_a = ws.cell(row=i+1, column=8)
            print('cell_a.value= ', cell_a.value)

            value1 = 0
            if cell_s.value != None:
               if 'ms' in str(cell_s.value):
                   value1 = float(cell_s.value.split('ms')[0])
               else:
                   value1 = float(cell_s.value)
            value2 = 0
            if cell_a.value != None:
               value2 = float(cell_a.value)
            if value2 - value1 < 0:
               print('pass -----------------------------')
               ws['I%s' % str(i+1)] = 'Pass' # I列标记Pass
    wb.save("基线表.xlsx")

read_excel_file("基线表.xlsx")
