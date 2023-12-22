# Copyright (c) 2023 Huawei Device Co., Ltd.
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import openpyxl
from openpyxl.styles import Alignment
import pandas as pd
import re

samples = []
components = []
file_path = 'D:/BenchMark_tool/BenchMark_js.txt'

workbook = openpyxl.Workbook()  # 创建工作簿
# worksheet = workbook.create_sheet(title='BenchMark')  # 创建工作表
worksheet = workbook.active  # 选择默认的表单
worksheet['A1'] = '页面名称'
worksheet['B1'] = 'js创建时间（单位ms）'

# 获取sample
def find_samples(file_path):
    with open(file_path, 'r') as file:
        lines = file.readlines()
        print("len(lines)：", len(lines))
        count = 0
        lineTime = 0
        for index, line in enumerate(lines):
            if 'please check Timer Name' in line:
                continue
            line_name = line.split(':')[3]
            if 'ms' in line:
               line_time = float(line.split(':')[4].split('ms')[0])
            else:
               line_time = float(line.split(':')[4].split('s')[0]) * 1000
            if index < len(lines)-1 and line_name == lines[index+1].split(':')[3]:
               count+=1
               lineTime += line_time
            if index < len(lines)-1 and line_name != lines[index+1].split(':')[3]:
               count+=1
               lineTime += line_time
               worksheet['A%s' % str(index+1)] = line_name 
               worksheet['B%s' % str(index+1)] = lineTime / count / 20
               count = 0
               lineTime = 0
            if index == len(lines)-1:
               count+=1
               lineTime += line_time
               worksheet['A%s' % str(index+1)] = line_name 
               worksheet['B%s' % str(index+1)] = lineTime / count / 20
               count = 0
               lineTime = 0
        workbook.save(filename='BenchMark_js.xlsx')  # 保存工作簿 

find_samples(file_path)



# 读取Excel文件
data = pd.read_excel('BenchMark_js.xlsx')
# 删除空行
data = data.dropna()
# 保存修改后的Excel文件
data.to_excel('BenchMark_js.xlsx', index=False)

wb = openpyxl.load_workbook('BenchMark_js.xlsx')
ws = wb.active  # 选择默认的表单
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 30
wb.save(filename='BenchMark_js.xlsx')  # 保存工作簿 