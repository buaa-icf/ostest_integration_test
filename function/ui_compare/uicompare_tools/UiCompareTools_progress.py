# -*- coding:utf-8 -*-

import os

import sys

import time

import json

import openpyxl

import threading

import subprocess

import filecmp

import shutil

from multiprocessing import Manager, Pool, Process

import numpy as np

from datetime import datetime

from PIL import Image

from PIL import ImageChops

from itertools import product



import argparse

parser = argparse.ArgumentParser(description='manual to this script')

parser.add_argument("--type", type=str)

parser.add_argument("--excel", type=str)

parser.add_argument("--dir", type=str)

parser.add_argument("--hap", type=str)

parser.add_argument("--json", type=str)

args = parser.parse_args()

mode = args.type

excel_path = args.excel

base_path = args.dir

hap_path = args.hap

json_path = args.json



new_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))

now_time = str(new_time)

dir_path = os.path.dirname(os.path.abspath(__file__))

reports_path = os.path.join(dir_path, "reports")

html_path = "./mode.html"

run_json_path = os.path.join(reports_path, "run_json_" + now_time)

run_report_path = os.path.join(reports_path, "run_report_{}.html".format(now_time))

diff_report_path = os.path.join(reports_path, "diff_report_{}.html".format(now_time))

device_list = []

image_size_diff = []



def get_file_list(file_path, file_key):

    # 返回指定文件夹file_path下所有的特定文件名列表

    file_list = []

    if os.path.isdir(file_path) == False:

        print("当前路径{}非文件夹，请重新输入正确的文件夹路径后重试".format(file_path))

        return []

    elif os.path.exists(file_path) == False:

        print("当前路径{}不存在，请重新输入正确的文件夹路径后重试".format(file_path))

        return []

    for file in os.listdir(file_path):

        if file.endswith(file_key):

            file_list.append(file)

    return file_list





def write_report_html(data, mode="0", image_path=""):

    if os.path.exists(html_path) == False:

        print("模板html不存在，未生成html报告")

        return

    if mode == "0":

        shutil.copy(html_path, run_report_path)

        module_data = ""

        total_data = "<table class='summary'>"

        print("report_data: ", data)

        for key, value in data.items():

            if key == "Total_data":

                with open(run_report_path, "a") as f:

                    f.write("""Save Image Path:</td>

            <td class="normal fourth">""" + image_path + """</td>

        </tr>

        <tr>

            <td class="normal first">Test Start/ End Time:</td>

            <td class="normal second">""" + value["start_end_time"] + """</td>

            <td class="normal third">Execution Time:</td>

            <td class="normal fourth">""" + value["run_time"] + """</td>

        </tr>

    </table>

    <table class="summary">

        <tr>

            <td class="normal modules">MOUDULES</td>

            <td class="normal total-tests">TOTAL_TESTS</td>

            <td class="normal passed">PASS</td>

            <td class="normal failed">CASE_FAIL</td>

            <td class="normal failed">GET_FAIL</td>

            <td class="normal passed">PASS_RATE</td>

        </tr>

        <tr>

            <th class="normal modules color-normal">""" + value["moudle_count"] + """</th>

            <th class="normal total-tests color-normal">""" + value["Total"] + """</th>

            <th class="normal passed color-passed">""" + value["Pass"] + """</th>

            <th class="normal failed color-failed">""" + value["Case_fail"] + """</th>

            <th class="normal failed color-failed">""" + value["Get_fail"] + """</th>

            <th class="normal passed color-passed">""" + value["Pass_rate"] + """</th>

        </tr>

    </table>

<!--{suites.context}-->

""")

            else:

                if isinstance(value, int):

                    continue

                total_data += """

        <tr>

            <th class="normal modules color-normal">""" + key + """</th>

            <th class="normal total-tests color-normal">""" + str(value["Total"]) + """</th>

            <th class="normal passed color-passed">""" + str(value["Pass"]) + """</th>

            <th class="normal failed color-failed">""" + str(value["Case_fail"]) + """</th>

            <th class="normal failed color-failed">""" + str(value["Get_fail"]) + """</th>

            <th class="normal passed color-passed">""" + value["Pass_rate"] + """</th>

        </tr>

"""

                title_flag = True

                for case_name, run_result, use_time in value["result_list"]:

                    if run_result == "PASS":

                        # 仅展示执行失败用例，执行成功的用例不展示

                        continue

                    elif title_flag == True:

                        module_data += """

                    <table class='test-suite'>

                        <tr>

                          <th class='title' colspan='4'>

                            <span class='title' onclick="document.all.""" + key + """.style.display=(document.all.""" + key + """.style.display =='none')?'':'none'">""" + key + """</span>

                          </th>

                        </tr>

                    </table>

                    <table class='test-suite' id='""" + key + """' style="display:''">

                        <tr>

                          <th class='normal module'>Module</th>

                          <th class='normal test-suite'>Testsuite</th>

                          <th class='normal test'>Testcase</th>

                          <th class='normal result'>Result</th>

                        </tr>

                        """

                        title_flag = False

                    module_data += """<tr>

                      <td class='normal module'>""" + key + """</td>

                      <td class='normal test-suite'>""" + value["module_name"] + """</td>

                      <td class='normal test'>""" + case_name + """</td>

                      <td class='normal result'>""" + run_result + """</td>

                    </tr>"""

                if title_flag == False:

                    module_data += """

</table>"""

        module_data += """

<!--{failures.context}-->

</div>

</body>

</html>"""

        total_data += "</table>"

        with open(run_report_path, "a") as f:

            f.write(total_data + module_data)

        print("获取图片报告已生成：", run_report_path)

    elif mode == "1":

        shutil.copy(html_path, diff_report_path)

        module_data = ""

        total_data = "<table class='summary'>"

        for key, value in data.items():

            if key == "Total_data":

                with open(diff_report_path, "a") as f:

                    f.write("""Diff Image Path:</td>

                    <td class="normal fourth">""" + image_path + """</td>

                </tr>

                <tr>

                    <td class="normal first">Test Start/ End Time:</td>

                    <td class="normal second">""" + value["start_end_time"] + """</td>

                    <td class="normal third">Execution Time:</td>

                    <td class="normal fourth">""" + value["run_time"] + """</td>

                </tr>

            </table>

            <table class="summary">

                <tr>

                    <td class="normal modules">MOUDULES</td>

                    <td class="normal total-tests">TOTAL_TESTS</td>

                    <td class="normal passed">PASS</td>

                    <td class="normal failed">CASE_FAIL</td>

                    <td class="normal failed">GET_FAIL</td>

                    <td class="normal failed">BASE_IMAGE</td>

                    <td class="normal failed">COMPARE_FAIL</td>

                    <td class="normal passed">PASS_RATE</td>

                </tr>

                <tr>

                    <th class="normal modules color-normal">""" + value["moudle_count"] + """</th>

                    <th class="normal total-tests color-normal">""" + value["Total"] + """</th>

                    <th class="normal passed color-passed">""" + value["Pass"] + """</th>

                    <th class="normal failed color-failed">""" + value["Case_fail"] + """</th>

                    <th class="normal failed color-failed">""" + value["Get_fail"] + """</th>

                    <th class="normal failed color-failed">""" + value["Base_image_fail"] + """</th>

                    <th class="normal failed color-failed">""" + value["Compare_fail"] + """</th>

                    <th class="normal passed color-passed">""" + value["Pass_rate"] + """</th>

                </tr>

            </table>

        <!--{suites.context}-->""")

            else:

                if isinstance(value, int):

                    continue

                total_data += """

                <tr>

                    <th class="normal modules color-normal">""" + key + """</th>

                    <th class="normal total-tests color-normal">""" + str(value["Total"]) + """</th>

                    <th class="normal passed color-passed">""" + str(value["Pass"]) + """</th>

                    <th class="normal failed color-failed">""" + str(value["Case_fail"]) + """</th>

                    <th class="normal failed color-failed">""" + str(value["Get_fail"]) + """</th>

                    <th class="normal failed color-failed">""" + str(value["Base_image_fail"]) + """</th>

                    <th class="normal failed color-failed">""" + str(value["Compare_fail"]) + """</th>

                    <th class="normal passed color-passed">""" + value["Pass_rate"] + """</th>

                </tr>

"""

                add_flag = True

                for case_name, run_result, use_time in value["result_list"]:

                    if run_result == "PASS":

                        # 仅展示对比失败用例，对比图片相同的用例不展示

                        continue

                    else:

                        if add_flag == True:

                            # 仅在有对比失败用例时，才写入标题栏

                            module_data += """

       <table class='test-suite'>

            <tr>

              <th class='title' colspan='4'>

                <span class='title' onclick="document.all.""" + key + """.style.display=(document.all.""" + key + """.style.display =='none')?'':'none'">""" + key + """</span>

              </th>

            </tr>

        </table>

        <table class='test-suite' id='""" + key + """' style="display:''">

            <tr>

              <th class='normal module'>Module</th>

              <th class='normal test'>Testcase</th>

              <th class='normal result'>Result</th>

            </tr>

        """

                            add_flag = False

                    module_data += """<tr>

          <td class='normal module'>""" + key + """</td>

          <td class='normal test'>""" + case_name + """</td>

        <td class='normal result'>""" + run_result + """</td>

        </tr>"""

                if add_flag == False:

                    module_data += """

        </table>"""

        module_data += """

        <!--{failures.context}-->

        </div>

        </body>

        </html>"""

        total_data += "</table>"

        with open(diff_report_path, "a") as f:

            f.write(total_data + module_data)

        print("对比图片报告已生成：", diff_report_path)





def run_cmd(command):

    # 执行cmd命令，返回执行日志

    p = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE)

    out, err = p.communicate()

    # print("out: ", out)

    # print("err: ", err)

    if err is None:

        return out.decode('utf-8')

    return err.decode('utf-8')





def checkResult(command, keyword="", case_name=""):

    # 检查设备是否连接成功或用例执行成功

    start_time = datetime.timestamp(datetime.now())

    out = run_cmd(command)

    end_time = datetime.timestamp(datetime.now())

    use_time = round(end_time - start_time, 3)

    if "#" not in command:

        # 批量执行，计算平均耗时

        use_time = int(use_time / len(out.splitlines()))

    print("{0} run use_time: {1}s".format(case_name, use_time))

    if "error: get bundle info failed" in out:

        print("【ERROR】测试应用hap未安装，请成功安装后重试\n", out)

        return "测试hap未安装"

    elif keyword in command:

        # 检查设备是否连接成功

        if "[Empty]" in out:

            # 设备未连接时，显示“[Empty]”

            print("设备未连接，请连接设备或检查线路连接是否正常，UI工具将开始尝试重连")

            return case_name + " Devices connect fail"

        elif out == "":

            # 无hdc.exe

            return "不支持hdc命令, 仅支持hdc_std命令"

        elif len(out) > 15:

            return "设备已连接，可正常使用"

    elif "execute timeout" in out:

        return "timeout"

    elif "Not found entry_test/ets/" in out:

        return True

    elif "stream=Tests run: 0" in out:

        return False

    elif keyword != "list targets":

        image_name_list = []

        error_case_list = []

        command_list = []

        # 获取用例名和失败用例名

        index_flag = 0

        for line_data in out.splitlines():

            # print(index_flag, line_data)

            # 用例执行成功

            if "OHOS_REPORT_STATUS_CODE: 0" in line_data:

                # 当前用例名

                case_name = out.splitlines()[index_flag - 1].split("=")[1]

                image_name_list.append([case_name, "PASS", use_time])

                print(case_name, " run succ !!!\n")

            elif "OHOS_REPORT_STATUS_CODE: -1" in line_data or "TestFinished-ResultCode: -1" in line_data:

                # 用例执行失败

                if "TestFinished-ResultCode: -1" in line_data and index_flag >= 3:

                    # 当前用例名

                    case_name = out.splitlines()[index_flag - 3].split("test=")[1]

                    error_info = out.splitlines()[index_flag + 1].split(":")[1]

                    print(out.splitlines()[index_flag + 1])

                    image_name_list.append([case_name, "CASE_FAIL: " + error_info, use_time])

                elif "OHOS_REPORT_STATUS_CODE: -1" in line_data and index_flag >= 2:

                    case_name = out.splitlines()[index_flag - 1].split("test=")[1]

                    # 执行失败原因

                    error_info = out.splitlines()[index_flag - 2]

                    print(case_name, " CASE_FAIL: " + error_info)

                    image_name_list.append([case_name, "CASE_FAIL: " + error_info, use_time])

                if "#" in command:

                    # 单条用例

                    error_case_list.append((command, case_name))

                elif command not in command_list:

                    # 多条用例，只添加一次即可（因为多条用例无法执行单用例）

                    error_case_list.append((command, case_name))

                    # 仅用来判断多用例执行失败情况

                    command_list.append(command)

                print(case_name, " run fail !!!\n")

            elif keyword in line_data:

                # 解析用例执行总结果

                print(line_data)

            index_flag += 1

        # print("image_name_list: ", image_name_list)

        return image_name_list, error_case_list

    else:

        return False





def delete_pictures(bundle_name):

    # 删除设备中应用原有截图

    app_list = run_cmd("hdc shell cd /data/app/el2/100/base/;ls")

    if bundle_name in app_list.split("\r\n"):

        # 删除指定应用文件夹内截图

        app_path = "/data/app/el2/100/base/{}/haps/entry_test/files/".format(bundle_name)

        print(app_path)

        run_cmd("hdc shell rm -rf {}*".format(app_path))

        time.sleep(1)

        out = run_cmd("hdc shell cd {};du -sk".format(app_path)).split(".")[0].strip()

        print(out)

        if "/bin/sh" in out:

            print("设备中不存在待删除图片!!!")

            return

        print("当前{0}文件夹内存（KB）：{1}".format(bundle_name, out))

        count = 0

        while int(out) > 10 and count <= 30:

            out = run_cmd("hdc shell cd {};du -sk".format(app_path)).split(".")[0].strip()

            print("当前{0}文件夹内存（KB）：{1}".format(bundle_name, out))

            count += 1

            time.sleep(5)

        print("设备中原有用例截图已全部删除!!!")

    else:

        print("设备中不存在待删除图片!!!")





class UiCompareTools():

    """

    UI对比工具

    """

    def run_testcases(self, sn, data, base_path, excel_path, report_path, excel_name, mode='0'):

        # 遍历执行Execl用例，并将设备端webp文件保存至本地

        # 设备解锁：兼容部分版本执行完用例设备会锁屏

        # run_cmd('hdc -t {} shell uinput -T -m 1280 800 1280 800 1000'.format(sn))

        # run_cmd('hdc -t {} shell uinput -T -m 500 2000 500 500 600'.format(sn))

        print("excel_path: ", excel_path)

        install_flag = False

        bundle_name = ""

        class_name = ""

        result_list = []

        pass_count = 0

        case_fail = 0

        get_fail = 0

        run_flag = False

        for case in data:

            try:

                class_name = case[0]

                case_name = case[1]

                bundle_name = case[2]

                if case_name is None:

                    run_flag = True

                else:

                    run_flag = False

                # 先卸载该应用，避免设备上存在已安装同名应用导致测试hap无法安装

                if install_flag == False:

                    # 卸载、安装仅在执行第一条用例前执行一次

                    install_flag = True

                    # 卸载hap

                    if bundle_name is not None:

                        run_cmd("hdc -t {0} uninstall {1}".format(sn, bundle_name))

                    # 安装hap

                    hap_install_path = os.path.join(hap_path, excel_name + ".hap")

                    print("hap_install_path: ", hap_install_path)

                    if os.path.exists(hap_install_path) == False:

                        # hap包路径不存在

                        for case in data:

                            case_fail += 1

                            result_list.append(

                                [case[1], "CASE_FAIL: The HAP path dose not exist: " + hap_install_path, "0"])

                        print("{}组件hap包{}安装路径不存在：{}".format(excel_name, bundle_name, hap_install_path))

                        break

                    out = run_cmd("hdc -t {0} install {1}".format(sn, hap_install_path))

                    print("Install result: ", out)

                    if "install failed" in out:

                        # hap包安装失败

                        for case in data:

                            case_fail += 1

                            result_list.append(

                                [case[1], "CASE_FAIL: Failed to install the HAP: " + hap_install_path, "0"])

                        print("{}组件hap包{}安装失败: {}".format(excel_name, bundle_name, hap_install_path))

                        break



                # 单条执行

                command = "hdc -t {3} shell aa test -b {0} -m entry_test -s unittest /ets/testrunner/OpenHarmonyTestRunner -s class {1}#{2} -w 180000 -s timeout 15000".format(case[2], case[0], case[1], sn)

                # 批量执行

                if case_name is None:

                    run_flag = True

                    # 删除设备中图片

                    delete_pictures(bundle_name)

                    command = "hdc -t {2} shell aa test -b {0} -m entry_test -s unittest /ets/testrunner/OpenHarmonyTestRunner -s class {1} -w 180000 -s timeout 15000".format(case[2], case[0], sn)

                result = checkResult(command, "stream=Tests run:", case_name)



                if result == True:

                    command = command.replace("/ets/testrunner/OpenHarmonyTestRunner", "OpenHarmonyTestRunner")

                    print("new command1: ", command)

                    result = checkResult(command, "stream=Tests run:", case_name)



                if result == True:

                    command = command.replace("OpenHarmonyTestRunner", "/ets/Testrunner/OpenHarmonyTestRunner")

                    print("new command2: ", command)

                    result = checkResult(command, "stream=Tests run:", case_name)



                count = 1

                while result == "timeout" and count < 6:

                    # 若用例执行超时，延长超时的时间5秒，最长延长至30秒，若仍超时，则判定用例执行失败

                    command = command.replace(str(15000 + 5000 * (count - 1)), str(15000 + 5000 * count))

                    print("new command3: ", command)

                    result = checkResult(command, "stream=Tests run:", case_name)

                    count += 1



                # 存在失败用例，重新执行

                count_error = 1

                while run_flag == False and count_error <= 5:

                    # 若执行失败，重试5次

                    if type(result) == tuple:

                        if len(result) == 2 and len(result[1]) >= 1:

                            print("重新执行失败用例：", result[1])

                            with open(report_path, "a") as file:

                                for error_info in result[1]:

                                    file.write(class_name + ": " + error_info[1] + " run fail\n")

                            # if run_flag == True:

                            #     # 删除设备中图片

                            #     delete_pictures(bundle_name)

                            result = checkResult(command, "stream=Tests run:", case_name)

                            count_error += 1

                        else:

                            break

                    else:

                        break



                if result == True or result == False:

                    print("用例执行失败：{}".format(command))

                    result = [[[case_name, "CASE_FAIL: testcase not run", "0"]], []]

                elif result == "测试hap未安装":

                    result = [[[case_name, "CASE_FAIL: Failed to install the HAP or The bundle name in the Excel is inconsistent with that in the HAP", "0"]], []]



                # 批量循环取图

                if class_name and (case_name is None):

                    row_num = 1

                    data_list = []

                    for case_name1, case_result, run_time in result[0]:

                        # 用例执行成功，将设备中webp文件保存至本地

                        if case_result == "PASS":

                            # 用例执行成功，才会创建保存图片的文件夹

                            if os.path.exists(base_path) == False:

                                os.makedirs(base_path)

                            print('base_path: ', base_path)

                            webp_path = os.path.join(base_path, "IMG_%s.webp" % case_name1)

                            device_webp = "data/app/el2/100/base/{0}/haps/entry_test/files/".format(case[2])

                            device_path = device_webp + "IMG_%s.webp" % case_name1

                            out_content2 = run_cmd("hdc -t %s file recv %s %s" % (sn, device_path, webp_path))

                            print(out_content2)

                            if "FileTransfer finish" in out_content2:

                                pass_count += 1

                                print("%s.webp文件保存成功\n" % case_name1)

                            elif "no such file" in out_content2:

                                get_fail += 1

                                file_path = out_content2.split("path:")[1]

                                # 取图失败，更新用例执行结果数据

                                result[0][row_num - 1][1] = "GET_FAIL: save image fail: {}".format(file_path)

                                print("设备中 %s 文件不存在，保存本地失败\n" % file_path)

                                with open(report_path, "a") as file:

                                    file.write("%s: %s save fail: 设备中该文件不存在\n" % (class_name, case_name1))

                            else:

                                count = 0

                                while count <= 30:

                                    # 增加延时，确保webp文件成功保存至本地

                                    time.sleep(1)

                                    count += 1

                                    if "FileTransfer finish" in out_content2:

                                        break

                        else:

                            case_fail += 1

                        # 数据存储至列表

                        data_list.append([row_num, class_name, case_name1, case[2], case_result])

                        row_num += 1

                    # 取完图后，删除设备中所有图片

                    delete_pictures(bundle_name)

                    try:

                        # 将用例执行结果写入Excel

                        self.write_excel_data(excel_path, sheet_name=case[0], result_data=data_list)

                    except Exception as e:

                        print("【ERROR】write_excel_data: %s" % str(e))



                # 执行单条取图

                elif class_name and case_name:

                    # 用例执行成功才从设备中取资源保存至本地

                    # 将用例执行结果写入Excel

                    # try:

                    #     self.write_excel_data(excel_path, result_data=result[0], compare=2)

                    # except Exception as e:

                    #     print("【ERROR】write_excel_data: %s" % str(e))



                    run_result = ""

                    if type(result) == tuple:

                        try:

                            run_result = result[0][0][1]

                        except:

                            result[0][0] = [case_name, "CASE_FAIL: run testcase fail", "0"]

                            run_result = result[0][0][1]

                    # print("run_result: ", run_result)

                    # 若用例执行成功，将设备中webp文件保存至本地

                    if run_result == "PASS":

                        # 用例执行成功，才会创建保存图片的文件夹

                        if os.path.exists(base_path) == False:

                            os.makedirs(base_path)

                        # run_count = 2

                        # while run_count <= 2:

                        # txt_path = os.path.join(base_path, case_name + ".txt")

                        # remote_txt = "/storage/media/100/local/files/Documents/TXT_test.txt"

                        # out_content2 = run_cmd("hdc -t %s file recv %s %s" % (sn, remote_txt, txt_path))

                        webp_path = os.path.join(base_path, case_name + ".webp")

                        remote_webp = "data/app/el2/100/base/{0}/haps/entry_test/files/IMG_test.webp".format(case[2])

                        out_content2 = run_cmd("hdc -t %s file recv %s %s" % (sn, remote_webp, webp_path))

                        print(out_content2)

                        if "FileTransfer finish" in out_content2:

                            # 累加执行用例成功且成功取图的用例个数

                            pass_count += 1

                            print("%s: %s.webp文件保存成功\n" % (sn, case_name))

                        elif "no such file" in out_content2:

                            # 取图失败，重新执行用例后再取图三次

                            print("设备中 %s 文件不存在，保存本地失败\n" % remote_webp)

                            result = checkResult(command, "stream=Tests run:", case_name)

                            if result == "PASS":

                                # 用例执行成功再取图片

                                count = 0

                                while count < 3:

                                    time.sleep(5)

                                    out_content2 = run_cmd("hdc -t %s file recv %s %s" % (sn, remote_webp, webp_path))

                                    print(out_content2)

                                    if "FileTransfer finish" in out_content2:

                                        pass_count += 1

                                        print("%s: %s.webp文件保存成功\n" % (sn, case_name))

                                        break

                                    count += 1

                            else:

                                # 用例执行失败

                                case_fail += 1

                        else:

                            count = 0

                            while count <= 30:

                                # 增加延时，确保webp文件成功保存至本地

                                time.sleep(1)

                                count += 1

                                if "FileTransfer finish" in out_content2:

                                    pass_count += 1

                                    print("%s: %s.webp文件保存成功\n" % (sn, case_name))

                                    break

                        if "no such file" in out_content2:

                            get_fail += 1

                            file_path = out_content2.split("path:")[1]

                            # 取图失败，更新用例执行结果数据

                            result[0][0][1] = "GET_FAIL: save image fail: {}".format(file_path)

                            print("设备%s中 %s 文件不存在，保存本地失败\n" % (sn, file_path))

                            with open(report_path, "a") as file:

                                file.write("%s: %s save fail: 设备%s中该文件不存在\n" % (class_name, case_name, sn))

                    else:

                        case_fail += 1

                    if len(result[0]) == 1:

                        result_list.append(result[0][0])

            except Exception as e:

                print("[ERROR] Get Image fail: ", str(e))

                time.sleep(5)

        # 用例执行结束，卸载hap

        if bundle_name is not None:

            run_cmd("hdc -t %s uninstall %s" % (sn, bundle_name))

        time.sleep(2)

        if mode == "0":

            print("{}.xlsx 中所有用例均已执行，基线图源均已获取".format(excel_name))

        elif mode == "1":

            print("{}.xlsx 中所有用例均已执行，对比图源均已获取".format(excel_name))

        if pass_count != len(result_list) - case_fail - get_fail:

            pass_count = len(result_list) - case_fail - get_fail

        if len(result_list) <= 0:

            total_num = case_fail + get_fail

        else:

            total_num = len(result_list)

        result_data = {"Total": total_num,

                       "Pass": pass_count,

                       "Case_fail": case_fail,

                       "Get_fail": get_fail,

                       "Pass_rate": str(int((pass_count/len(data))*100)) + "%",

                       "result_list": result_list,

                       "module_name": class_name}

        # print("result_data: ", result_data)

        return result_data



    def read_excel_data(self, excel_path):

        # 读取Excel表中类名和用例名

        if os.path.exists(excel_path) == False:

            print("当前Excel路径不存在: %s" % excel_path)

            return False

        else:

            data_list = []

            try:

                wb = openpyxl.load_workbook(excel_path)

                # 默认取第一张表

                sheet = wb.worksheets[0]

                rows = sheet.max_row



                for row in range(1, rows):

                    # 获取第二列的数据:类名

                    class_name = sheet.cell(row + 1, 2).value

                    if class_name is None:

                        print("Excel表中有效数据行数: ", row - 1)

                        break

                    # 获取第三列的数据：用例名

                    case_name = sheet.cell(row + 1, 3).value

                    # 获取第四列的数据：bundle_name

                    bundle_name = sheet.cell(row + 1, 4).value

                    data_list.append((class_name, case_name, bundle_name))

                print('data_list: ', data_list)

            except Exception as e:

                print(str(e))

            return data_list



    def write_excel_data(self, report_path, sheet_name="Sheet1", result_data=[], x=1, compare=0):

        # 将执行报告数据写入Excel表中

        if os.path.exists(excel_path) == False:

            print("当前Excel路径不存在: %s" % excel_path)

            return False

        print("write_excel_data: ", excel_path)

        # 读取Excel表数据

        try:

            wb = openpyxl.load_workbook(excel_path)

            if compare == 2:

                # 默认取第一张表(执行单条用例)

                sheet = wb.worksheets[0]

            elif sheet_name in wb.sheetnames:

                sheet = wb[sheet_name]

            else:

                # 若sheet页不存在，创建新的sheet页

                sheet = wb.create_sheet(sheet_name)

                print("add sheet %s succ" % sheet_name)

                if compare == 0:

                    sheet.append(["序号", "类名", "用例名", "bundle name", "生成图片结果"])

                    print("标题写入完成")

                elif compare == 1:

                    sheet.append(["用例名", "对比结果", "耗时"])

                    for testcase in result_data:

                        sheet.append(testcase)

                    print("对比结果写入完成")

            wb.save(excel_path)

            # 写入result数据

            rows = sheet.max_row

            cols = sheet.max_column

            if compare == 0 or compare == 2:

                if result_data == []:

                    print("[ERROR] write excel data error")

                elif len(result_data) == 1:

                    # 单条用例

                    for num in range(1, rows + 1):

                        # print("num: ", num)

                        if sheet.cell(num, 3).value == result_data[0][0]:

                            for i in range(5, 100):

                                if sheet.cell(num, i).value is None:

                                    sheet.cell(num, i, result_data[0][1])

                                    if num == 2:

                                        print("单条用例：报告时间写入完成")

                                        sheet.cell(1, i, now_time)

                                    break

                            break

                else:

                    # 多条用例

                    if sheet.cell(x + 1, 1).value is None:

                        # 首次写入

                        for data in result_data:

                            sheet.append([data[0], data[1], data[2], data[3], data[4]])

                            print("write %s succ" % data[2])

                    else:

                        # 非首次写入

                        if sheet.cell(rows, cols + 1).value is None:

                            sheet.cell(1, cols + 1, now_time)

                            print("多条用例：报告时间写入完成")

                        cols += 1

                        for data in result_data:

                            sheet.cell(data[0] + 1, cols, data[4])

                            print("add write %s succ" % data[2])

            wb.save(excel_path)

            if sheet.cell(2, 1).value is None:

                # 删除空行

                sheet.delete_rows(2)

            # 保存Excel文件

            wb.save(excel_path)

            print('result_data: %s succ' % result_data)

        except Exception as e:

            with open(report_path, "a") as r1:

                r1.write("write Excel data error: {}\n".format(excel_path))

            print(str(e), type(e))

            if "Permission denied" in str(e):

                print("【ERROR】请先关闭本地已打开的Excel表格，再重新执行")



    def get_diff_files(self, excel_path, base_path, compare_path, run_result, now_time):

        # 比较两个文件夹中所有的webp文件，返回所有不同的文件名

        base_path_exists = True

        if os.path.isdir(base_path) == False:

            base_path_exists = False

            print("{}基线图源文件夹路径不存在".format(base_path))

        report_diff_list = []

        case_fail_count = 0

        get_fail_count = 0

        base_image_fail = 0

        compare_fail = 0

        pass_count = 0

        for case_name, case_result, use_time in run_result["result_list"]:

            if case_result == "PASS":

                baseImage_path = os.path.join(base_path, case_name + ".webp")

                compareImage_path = os.path.join(compare_path, case_name + ".webp")

                print("baseImage_path: ", baseImage_path)

                print("compareImage_path: ", compareImage_path)

                use_time = "0"

                if base_path_exists == False:

                    # 基线图源文件夹路径不存在

                    base_image_fail += 1

                    case_result = "BASE_IMAGE: BaseImage moudle path does not exist: {}".format(base_path)

                elif os.path.exists(baseImage_path) == False:

                    # 基线图源文件不存在

                    base_image_fail += 1

                    case_result = "BASE_IMAGE: BaseImage does not exist: {}".format(baseImage_path)

                elif os.path.exists(compareImage_path) == False:

                    # 对比图片文件不存在

                    get_fail_count += 1

                    case_result = "GET_FAIL: Save image fail: {}".format(compareImage_path)

                else:

                    start_time = datetime.timestamp(datetime.now())

                    # 对比图片，不同时会生成差异图

                    if self.compare_files(baseImage_path, compareImage_path) == False:

                        case_result = "COMPARE_FAIL"

                        compare_fail += 1

                    else:

                        case_result = "PASS"

                        pass_count += 1

                    end_time = datetime.timestamp(datetime.now())

                    use_time = str(round(end_time - start_time, 3))

            elif "CASE_FAIL" in case_result:

                case_fail_count += 1

            elif "GET_FAIL" in case_result:

                get_fail_count += 1

            # 将用例执行结果数据保存至列表

            report_diff_list.append([case_name, case_result, use_time])



        # 将对比执行结果写入Excel

        # self.write_excel_data(excel_path, sheet_name="diff_%s" % now_time, result_data=report_diff_list, compare=1)

        case_count = run_result["Total"]

        if case_count == 0:

            pass_rate = 0

        else:

            # 防止除数为0，导致异常

            pass_rate = int(pass_count / case_count * 100)

        if pass_count != case_count - case_fail_count - get_fail_count - base_image_fail - compare_fail:

            pass_count = case_count - case_fail_count - get_fail_count - base_image_fail - compare_fail

        result_data = {"Total": case_count,

                       "Pass": pass_count,

                       "Case_fail": case_fail_count,

                       "Get_fail": get_fail_count,

                       "Base_image_fail": base_image_fail,

                       "Compare_fail": compare_fail,

                       "Pass_rate": str(pass_rate) + "%",

                       "result_list": report_diff_list}

        print(result_data)

        return result_data



    def modify_image_color(self, num, height_unit, width, img_array, dst):

        # 修改图片像素颜色

        height_1 = num * height_unit

        height_2 = (num + 1) * height_unit

        for h, w in product(range(height_1, height_2), range(width)):

            (b, g, r) = img_array[h, w]

            if (b, g, r) == (0, 0, 0):  # 黑色

                img_array[h, w] = (255, 255, 255)  # 白色

            else:

                img_array[h, w] = (255, 0, 0)  # 剩余差异部分红色显示

            dst[h, w] = img_array[h, w]

        return dst



    def compare_files(self, file1, file2):

        # 比较两个webp文件，如果它们相同则返回True，否则返回False

        if os.path.exists(file1) == False:

            print("当前文件路径不存在，请确认后重试：", file1)

            return False

        if os.path.exists(file2) == False:

            print("当前文件路径不存在，请确认后重试：", file2)

            return False

        # if filecmp.cmp(file1, file2) == True:

        if self.compare_file(file1, file2) == True:

            print("两张图片内容：完全相同")

            return True

        else:

            """

            比较图片，如果有不同则生成展示不同的图片

            @参数一: file1: 基线图片的路径

            @参数二: file2: 对比图片的路径

            """

            module_name = file1.split("\\")[-2]

            diff_path = os.path.join(reports_path, "diff_" + now_time, module_name)

            print("diff_path: ", diff_path)

            if os.path.exists(diff_path) == False:

                os.makedirs(diff_path)

            try:

                image_one = Image.open(file1)

                image_two = Image.open(file2)

            except Exception as e:

                print(str(e))

                return False

            size1 = image_one.size

            size2 = image_two.size

            print("BaseImage size: {}\nCompareImage size: {}".format(size1, size2))

            if size1 != size2:

                # 两张图片大小不一致, 无需对比，且不会生成差异图

                if module_name not in image_size_diff:

                    image_size_diff.append(module_name)

                path = os.path.join(reports_path, module_name + ".txt")

                with open(path, "a") as file:

                    file.write("截图大小不一致组件列表: %s !!!\n" % str(image_size_diff))

                with open(report_path, "a") as file:

                    file.write("截图大小不一致组件列表: %s !!!\n" % str(image_size_diff))

                return False

            try:

                diff = ImageChops.difference(image_one, image_two)

                if diff.getbbox() is None:

                    # 两张图片相同

                    return True

                img_array = np.array(diff)  # 把图像转成数组格式img = np.asarray(image)

                shape = img_array.shape

                height = shape[0]

                width = shape[1]

                dst = np.zeros((height, width, 3))

                # for h, w in product(range(height), range(width)):

                #     (b, g, r) = img_array[h, w]

                #     if (b, g, r) == (0, 0, 0):  # 黑色

                #         img_array[h, w] = (255, 255, 255)  # 白色

                #     else:

                #         img_array[h, w] = (255, 0, 0)  # 剩余差异部分红色显示

                #     dst[h, w] = img_array[h, w]

                black_pixels = np.where(

                    (img_array[:, :, 0] == 0) & (img_array[:, :, 1] == 0) & (img_array[:, :, 2] == 0))

                dst[black_pixels] = [255, 255, 255]

                other_pixels = np.where(

                    (img_array[:, :, 0] != 0) | (img_array[:, :, 1] != 0) | (img_array[:, :, 2] != 0))

                dst[other_pixels] = [255, 0, 0]

                diff_image = Image.fromarray(np.uint8(dst))

                print("【+】We are the not same!")

                # 保存差异图、基线图和对比图到diff_path下，命名规则：用例名_diff，用例名_base，用例名_compare

                case_name = file1.split("\\")[-1].split(".")[0]

                diff_save_path = os.path.join(diff_path, case_name + '_diff.webp')

                base_image_path = os.path.join(diff_path, case_name + '_base.webp')

                compare_image_path = os.path.join(diff_path, case_name + '_compare.webp')

                diff_image.save(diff_save_path)

                shutil.copy(file1, base_image_path)

                shutil.copy(file2, compare_image_path)

            except ValueError as e:

                text = ("表示图片大小和box对应的宽度不一致，参考API说明：Pastes another image into this image."

                        "The box argument is either a 2-tuple giving the upper left corner, a 4-tuple defining the left, upper, "

                        "right, and lower pixel coordinate, or None (same as (0, 0)). If a 4-tuple is given, the size of the pasted "

                        "image must match the size of the region.使用2纬的box避免上述问题")

                print("【{0}】{1}".format(e, text))

            print("两张图片内容：存在差异")

            return False



    def compare_file(self, file1, file2):

        with open(file1, "rb") as f1:

            data1 = f1.read()

        with open(file2, "rb") as f2:

            data2 = f2.read()

        return data1 == data2



    def generate_report(self, different_files, report_path):

        # 生成报告显示具体哪个webp文件不同

        if len(different_files) == 0:

            print("两个文件夹中的所有webp文件都相同。")

            return



        report = "以下图片在两个文件夹中存在差异：\n"

        for file in different_files:

            report += f"{file}\n"

        with open(report_path, "w") as r1:

            r1.write(report)

        print("webp文件比对完成，已生成报告：", report_path)





def connect_device(device_list):

    # 检查设备是否成功连接

    time.sleep(3)

    result = run_cmd("hdc list targets")

    now_device_list = result.split("\r\n")[-1]

    print("应连接设备 %d 台" % len(device_list))



    # 先循环等待是否全部连接成功

    count = 0

    while len(device_list) != len(now_device_list) and count < 5:

        time.sleep(2)

        count += 1



    # 获取未连接设备sn号

    disconnect_device = []

    if len(device_list) != len(now_device_list):

        for device in device_list:

            if device not in now_device_list:

                disconnect_device.append(device)

    else:

        print("设备均已成功连接PC")



    # 重试拉起未连接设备hdc，再查询

    if len(disconnect_device) >= 1:

        print("设备连接异常, 开始重连设备")

        for sn in disconnect_device:

            for i in range(1, 16):

                SN = run_cmd("hdc -t {} shell power-shell setmode 602".format(sn))

                time.sleep(2)

                # print("SN: ", SN)

                if "[Fail]" in SN:

                    print("重连第 %d 次" % i)

                    continue

                else:

                    break

            run_cmd("hdc -t {} shell power-shell display -s 0".format(sn))

        if len(device_list) == len(run_cmd("hdc list targets").split("\r\n")[-1]):

            print("设备均已成功连接PC")

        else:

            # 更新设备列表

            device_list = run_cmd("hdc list targets").split("\r\n")[-1]

            return device_list





def delete_and_reboot(sn):

    # 跑用例前删包，重启操作

    run_cmd("hdc -t {} shell mount -o remount,rw /sys_prod".format(sn))

    run_cmd("hdc -t {} shell rm -rf /sys_prod/app/SystemResourcesOverlay".format(sn))

    # 重启设备

    #run_cmd("hdc -t {} shell reboot".format(sn))

    #print("设备 {} 重启中，请耐心等待50秒......".format(sn))

    #time.sleep(50)

    # 当前设备类型

    device = run_cmd('hdc -t {} shell param get const.product.model'.format(sn))

    if "Fail" in device:

        # 设备断连，尝试重连

        connect_device(device_list)

    else:

        print("当前设备类型：", device)

    # 解锁设备

    if "WGR" in device:

        # 若是WGR需点击屏幕中间才可解锁

        run_cmd('hdc -t {} shell uinput -T -m 1280 800 1280 800 1000'.format(sn))

    elif "NOH" in device:

        # 手机：滑动界面即可解锁

        run_cmd('hdc -t {} shell uinput -T -m 500 2000 500 500 600'.format(sn))

    else:

        # 设备断连，尝试重连

        connect_device([sn])

        run_cmd('hdc -t {} shell uinput -T -m 1280 800 1280 800 1000'.format(sn))

        run_cmd('hdc -t {} shell uinput -T -m 500 2000 500 500 600'.format(sn))

    # 设置设备常亮

    out = run_cmd('hdc -t {} shell power-shell setmode 602'.format(sn))

    if "Mode Success" in out:

        print("当前设备 {} 已设置常亮！！！".format(sn))

        # 调低亮度，确保有足够电量进行测试

        run_cmd('hdc -t {} shell power-shell display -s 0'.format(sn))

    else:

        print("当前设备 {} 连接异常，请成功连接设备后重新执行".format(sn))





def run_get_image(args):

    # 获取基线图

    # device_dict, excel_path, excel_name, report_path, base_path, now_time, dir_path

    global device_dict

    device_dict = args[0]

    excel_path = args[1]

    excel_name = args[2]

    report_path = args[3]

    now_time = args[5]

    dir_path = args[6]

    print(excel_name, " before run device_dict: ", device_dict)

    if len(device_dict) == 1:

        for key, value in device_dict.items():

            sn = key

    elif len(device_dict) > 1:

        sn = device_dict.keys()[0]

    # print("sn: ", sn)

    run_flag = False

    if device_dict[sn] == True:

        # 有空闲设备时，占用并更新设备状态不可用（False）

        tmp = device_dict

        tmp[sn] = False

        device_dict = tmp

        print(excel_name, " running device_dict: ", device_dict)

    elif device_dict[sn] == False:

        index = 0

        # 单个组件执行用例超过50分钟，可能该设备已闲置，启用该设备

        while run_flag == False and index < 50 * 60:

            for device, status in device_dict.items():

                # 监测到有空闲设备时

                if status == True:

                    # 有空闲设备时，占用并更新设备状态不可用（False）

                    tmp = device_dict

                    tmp[device] = False

                    device_dict = tmp

                    sn = device

                    run_flag = True

                    print(excel_name, " running device_dict: ", device_dict)

                    break

            if run_flag == True:

                # 有空闲设备时，退出循环，开始执行用例

                break

            # 等待直到出现空闲设备

            time.sleep(1)

            index += 1

    global base_path

    base_path = os.path.join(dir_path, "BaseImages_" + now_time)

    print('base_path: ', base_path)

    # 实例化对象

    uiTools = UiCompareTools()

    data = uiTools.read_excel_data(excel_path)

    result_dict = {}

    if data != False:

        save_image_path = os.path.join(base_path, excel_name)

        try:

            case_result = uiTools.run_testcases(sn, data, save_image_path, excel_path, report_path, excel_name, mode="0")

            result_dict[excel_name] = case_result

            # 将本次执行结果保存至json文件中，确保遇到异常中断，仍可生成总的html报告

            if os.path.exists(run_json_path) == False:

                os.makedirs(run_json_path)

            json_data_path = os.path.join(run_json_path, excel_name + ".json")

            with open(json_data_path, "w") as f:

                json.dump(result_dict, f)

                # print("成功写入json文件：", result_dict)

        except Exception as e:

            print("[ERROR] get BaseImage fail: ", str(e))

        print("result_dict: ", result_dict)

        # 用例执行完，更新当前设备状态为可用（True）

        device_dict[sn] = True

        print(excel_name, " after run device_dict: ", device_dict)

    else:

        print("[ERROR] The Excel path does not exist !!!")

    return device_dict, result_dict





def run_compare_image(args):

    # 获取对比图，并对比图片

    # device_dict, excel_path, excel_name, report_path, base_path, now_time, dir_path

    global device_dict

    device_dict = args[0]

    excel_path = args[1]

    excel_name = args[2]

    report_path = args[3]

    base_path = args[4]

    now_time = args[5]

    dir_path = args[6]

    print(excel_name, " before run device_dict: ", device_dict)

    if len(device_dict) == 1:

        for key, value in device_dict.items():

            sn = key

    elif len(device_dict) > 1:

        sn = device_dict.keys()[0]

    # print("sn: ", sn)

    run_flag = False

    if device_dict[sn] == True:

        # 有空闲设备时，占用并更新设备状态不可用（False）

        tmp = device_dict

        tmp[sn] = False

        device_dict = tmp

        print(excel_name, " running device_dict: ", device_dict)

    elif device_dict[sn] == False:

        index = 0

        # 单个组件执行用例超过50分钟，可能该设备已闲置，启用该设备

        while run_flag == False and index < 50*60:

            for device, status in device_dict.items():

                # 监测到有空闲设备时

                if status == True:

                    # 有空闲设备时，占用并更新设备状态不可用（False）

                    tmp = device_dict

                    tmp[device] = False

                    device_dict = tmp

                    sn = device

                    run_flag = True

                    print(excel_name, " running device_dict: ", device_dict)

                    break

            if run_flag == True:

                # 有空闲设备时，退出循环，开始执行用例

                break

            # 等待直到出现空闲设备

            time.sleep(1)

            index += 1



    compare_path = os.path.join(dir_path, "CompareImages_" + now_time)

    print('compare_path: ', compare_path)

    count = 0

    # 校验基线图片路径是否合法

    while os.path.isdir(base_path) == False or os.path.exists(base_path) == False:

        if os.path.isdir(base_path) == False:

            base_path = input("路径非文件夹，请重新输入基线图源文件夹所在路径：")

        elif os.path.exists(base_path) == False:

            base_path = input("路径不存在，请重新输入基线图源文件夹所在路径：")

        count += 1

        if os.path.isdir(base_path) == True and os.path.exists(base_path) == True:

            # 路径输入正确，跳出循环继续往下执行

            break

        if count == 3:

            # 输入三次非法参数，程序退出

            sys.exit(0)



    # 实例化对象

    uiTools = UiCompareTools()

    data = uiTools.read_excel_data(excel_path)

    diff_dict = {}

    if data != False:

        # 基线图片路径

        base_image_path = os.path.join(base_path, excel_name)

        # 对比图片路径

        save_image_path = os.path.join(compare_path, excel_name)

        # 执行生成对比图片用例，并将执行和保存图片的结果写入Excel表中

        uiTools = UiCompareTools()

        try:

            case_result = uiTools.run_testcases(sn, data, save_image_path, excel_path, report_path, excel_name, mode="1")

            # 用例执行完，更新当前设备状态为可用（True）

            device_dict[sn] = True

            # 开始对比图片，并将对比结果写入Excel表中，有差异时生成差异图保存至reports\diff_XXX\module_name\XXX.webp

            different_files = uiTools.get_diff_files(excel_path, base_image_path, save_image_path, case_result, now_time)

            # 将执行用例的结果、保存图片的结果与对比图片的结果合并，得到最终展示至报告中的结果

            # 更新该模块用例总数

            if different_files["Total"] < len(different_files["result_list"]):

                different_files["Total"] = len(different_files["result_list"])

            # 将该模块数据添加至字典

            diff_dict[excel_name] = different_files

            # 将本次执行结果保存至json文件中，确保遇到异常中断，仍可生成总的html报告

            if os.path.exists(run_json_path) == False:

                os.makedirs(run_json_path)

            json_data_path = os.path.join(run_json_path, excel_name + "_diff.json")

            with open(json_data_path, "w") as f:

                json.dump(diff_dict, f)

                print("成功写入json文件：", result_dict)



            # 生成对比报告

            report_diff_path = os.path.join(reports_path, "report_diff_{0}.txt".format(now_time))

            uiTools.generate_report(different_files["result_list"], report_diff_path)

        except Exception as e:

            print("[ERROR] run Compare fail: ", str(e))

        print("{}.xlsx 中所有用例图片均已对比完成".format(excel_name))

        print("diff_dict: ", diff_dict)

    else:

        print("[ERROR] The Excel path does not exist !!!")

    print(excel_name, " after run device_dict: ", device_dict)

    return device_dict, diff_dict





if __name__ == '__main__':

    # 程序入口

    # 1小时后开始生成基线图

    # time.sleep(1 * 60 * 60)

    start_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))

    start_time_second = datetime.timestamp(datetime.now())

    # 创建reports文件夹

    if os.path.exists(reports_path) == False:

        os.makedirs(reports_path)

    result_list = []

    if mode == "0" or mode == "1":

        # 判断用例数据的Excel所在文件夹内是否包含有效数据（必须为xlsx格式的Excel）

        count = 0

        while os.path.exists(excel_path) == False:

            excel_path = input("路径不存在，请重新输入Excel所在文件夹路径：")

            count += 1

            if os.path.exists(excel_path) == False and count == 3:

                # 尝试三次错误输入后，退出程序

                sys.exit(0)



        # 获取所有Excel表的全路径

        excel_name_list = []

        excel_path_list = []

        if os.path.exists(excel_path) == True:

            excel_name_list = get_file_list(excel_path, ".xlsx")

            count = 0

            # 路径无效时，尝试三次输入

            while len(excel_name_list) == 0:

                excel_path = input("路径中不存在xlsx格式的Excel文件，请检查后重新输入Excel所在文件夹路径：")

                excel_name_list = get_file_list(excel_path, ".xlsx")

                count += 1

                if len(excel_name_list) == 0 and count == 3:

                    # 尝试三次错误输入后，退出程序

                    sys.exit(0)

            # 路径有效时，获取全路径列表

            for excel_name in excel_name_list:

                excel_full_path = os.path.join(excel_path, excel_name)

                excel_name = excel_name.split(".")[0]

                excel_path_list.append((excel_full_path, excel_name))



        flag = 0

        func_name = ""

        while flag < 3:

            if mode == "0":

                # 获取基线图

                func_name = run_get_image

                break

            elif mode == "1":

                # 对比图片

                func_name = run_compare_image

                break

            else:

                mode = input("您输入有误，请重新输入（输入0可获取基线图源，输入1可对比图片）:")

                flag += 1

                if flag == 3:

                    # 错输三次，退出程序

                    sys.exit(0)

        report_path = os.path.join(reports_path, "reports_run_%s.txt" % now_time)



        # 兼容本地无hdc.exe的情况

        content = run_cmd("hdc list targets")

        print("当前设备SN号: \n", content)

        if content == "":

            position = run_cmd("where hdc_std")

            print("本地hdc_std.exe所在路径: ", position)

            position = position.splitlines()[0]

            print("本地hdc_std.exe所在第一个路径: ", position)

            shutil.copy(position, position.replace("hdc_std.exe", "hdc.exe"))

            content = run_cmd("hdc list targets")



        device_list = content.split("\r\n")[:-1]

        print("device_list: ", device_list)

        # 最大进程数，即连接设备数

        process_num = len(device_list)

        # 跑用例前删包，重启操作

        if process_num == 1:

            # 单台设备

            delete_and_reboot(device_list[0])

            print("单设备执行")

        elif process_num > 1:

            # 多设备、多进程

            print("多设备、多进程执行")

            with Pool(process_num) as p:

                outputs = p.map(delete_and_reboot, device_list)

            # 关闭进程池

            p.terminate()

        # 检查设备是否成功连接，若连接异常尝试重连

        connect_device(device_list)



        # 利用进程池Pool自动帮我们管理子进程

        excel_num = len(excel_path_list)

        device_num = len(device_list)

        print("excel_path_list: ", excel_path_list)

        dir_path = os.path.join(dir_path)



        if mode == "0":

            base_path = ""



        if device_num == 1:

            # 仅连接单台设备时

            device_dict = {device_list[0]: True}

            for excel_path, excel_name in excel_path_list:

                result_data = []

                if mode == "0":

                    result_data = run_get_image((device_dict, excel_path, excel_name, report_path, base_path, now_time, dir_path))

                elif mode == "1":

                    result_data = run_compare_image(

                        (device_dict, excel_path, excel_name, report_path, base_path, now_time, dir_path))

                else:

                    print("您输入的type有误，请您重新输入！")

                    sys.exit(0)

                result_list.append(result_data)

        elif device_num > 1:

            # 连接多台设备时

            # 未执行用例前，所有设备均可用（True）

            device_dict = Manager().dict()

            for device in device_list:

                device_dict[device] = True

            # 最大进程数：PC已连接的设备数

            p = Pool(device_num)

            res_l = []

            excel_index = 0

            for excel_path, excel_name in excel_path_list:

                res = p.apply_async(func_name, ((device_dict, excel_path, excel_name, report_path, base_path, now_time, dir_path),))

                res_l.append(res)

                excel_index += 1

            p.close()

            p.join()



            for res in res_l:

                # 汇总用例执行结果

                result_list.append(res.get())

            # 关闭进程池

            p.terminate()

    if mode == "2":

        # 获取所有的json路径

        all_json_path = []

        json_name_list = get_file_list(json_path, ".json")

        if "_diff.json" in json_name_list[0]:

            mode = "1"

        else:

            mode = "0"

            base_path = ""

        for json_name in json_name_list:

            all_json_path.append(os.path.join(json_path, json_name))

        print("all_json_path: ", all_json_path)

        for json_data_path in all_json_path:

            with open(json_data_path, "r") as f:

                data = json.loads(f.read())

                # print("解析后json文件：", data)

                result_list.append(({}, data))



    end_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))

    end_time_second = datetime.timestamp(datetime.now())

    run_time = end_time_second - start_time_second

    moudle_count = len(result_list)

    total_count = 0

    pass_count = 0

    case_fail_count = 0

    get_fail_count = 0



    if mode == "0":

        # 获取基线图

        result_dict = {}

        for run_dict in result_list:

            for key, value in run_dict[1].items():

                result_dict[key] = value

                total_count += value["Total"]

                pass_count += value["Pass"]

                case_fail_count += value["Case_fail"]

                get_fail_count += value["Get_fail"]

        if total_count == 0:

            pass_rate = 0

        else:

            pass_rate = int(pass_count / total_count * 100)

        # if pass_count != total_count - case_fail_count - get_fail_count:

        #     pass_count = total_count - case_fail_count - get_fail_count

        result_dict["Total_data"] = {"moudle_count": str(moudle_count),

                                     "Total": str(total_count),

                                     "Pass": str(pass_count),

                                     "Case_fail": str(case_fail_count),

                                     "Get_fail": str(get_fail_count),

                                     "Pass_rate": str(pass_rate) + "%",

                                     "start_end_time": str(start_time) + "/ " + str(end_time),

                                     "run_time": str(time.strftime("%H:%M:%S", time.gmtime(run_time)))

                                     }

        # 生成获取基线图的html报告

        write_report_html(result_dict, mode="0", image_path=base_path)

        print("获取图片结束，result_dict: ", result_dict)

    elif mode == "1":

        # 与基线图对比，并生成差异图和报告

        base_image_fail_count = 0

        compare_fail_count = 0

        diff_dict = {}

        for run_dict in result_list:

            for key, value in run_dict[1].items():

                diff_dict[key] = value

                # 累加获取各状态总数

                total_count += value["Total"]

                pass_count += value["Pass"]

                case_fail_count += value["Case_fail"]

                get_fail_count += value["Get_fail"]

                base_image_fail_count += value["Base_image_fail"]

                compare_fail_count += value["Compare_fail"]

        if total_count == 0:

            pass_rate = 0

        else:

            pass_rate = int(pass_count / total_count * 100)

        diff_dict["Total_data"] = {"moudle_count": str(moudle_count),

                                   "Total": str(total_count),

                                   "Pass": str(pass_count),

                                   "Case_fail": str(case_fail_count),

                                   "Get_fail": str(get_fail_count),

                                   "Base_image_fail": str(base_image_fail_count),

                                   "Compare_fail": str(compare_fail_count),

                                   "Pass_rate": str(pass_rate) + "%",

                                   "start_end_time": str(start_time) + "/ " + str(end_time),

                                   "run_time": str(time.strftime("%H:%M:%S", time.gmtime(run_time)))}

        diff_path = os.path.join(reports_path, "diff_" + str(now_time))

        # 生成对比图片后的html报告

        write_report_html(diff_dict, mode="1", image_path=diff_path)

        print("对比图片结束，diff_dict: ", diff_dict)

